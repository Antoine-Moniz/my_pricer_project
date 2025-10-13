import os
import pandas as pd
from openpyxl import Workbook
import numpy as np
import streamlit as st
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from functools import lru_cache


class TreeExporter:
    """Classe responsable des exports Excel et preview de l'arbre."""

    # ------------------------------
    # Helpers
    # ------------------------------
    @staticmethod
    def _prepare_workbook():
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        return wb

    @staticmethod
    def _fill_field_sheet(ws, node, steps, attr, offset):
        """Remplit une feuille Excel avec les valeurs d'un attribut de l'arbre."""
        for col in range(1, steps + 2):
            ws.cell(row=offset + steps + 1, column=col, value=getattr(node, attr))
            up, down = node.upper_node, node.lower_node
            for row in range(2, steps + 2):
                if up:
                    ws.cell(row=offset + steps - row + 2, column=col, value=getattr(up, attr))
                    up = up.upper_node
                if down:
                    ws.cell(row=offset + steps + row, column=col, value=getattr(down, attr))
                    down = down.lower_node
            node = node.next_mid

    @staticmethod
    def _build_preview(tree, preview_steps):
        """Construit le DataFrame preview (spots + valorisation)."""
        data_preview, node = [], tree.root
        for step in range(min(preview_steps, tree.params.nb_steps)):
            data_preview.append({
                "step": step,
                "spot": getattr(node, "spot", None),
                "valorisation": getattr(node, "opt_val", None)
            })
            node = node.next_mid if node and node.next_mid else None
            if node is None:
                break
        return pd.DataFrame(data_preview)

    # ------------------------------
    # Public
    # ------------------------------
    @staticmethod
    def save_to_excel(tree, fields, filename, offset=0, out_dir="exports", preview_steps=5):
        """Sauvegarde l'arbre trinomial dans un fichier Excel et retourne chemin + DataFrame preview."""
        if tree.last is None or getattr(tree.root, "opt_val", None) is None:
            tree.price_backward()

        base_dir = os.path.dirname(os.path.abspath(__file__))
        out_path = os.path.join(base_dir, out_dir)
        os.makedirs(out_path, exist_ok=True)

        rename_map = {"valorisation": "opt_val"}
        wb = TreeExporter._prepare_workbook()

        for field in fields:
            attr = rename_map.get(field, field)
            ws, node, steps = wb.create_sheet(f"arbre_{field}"), tree.root, tree.params.nb_steps
            TreeExporter._fill_field_sheet(ws, node, steps, attr, offset)

        file_path = os.path.join(out_path, f"{filename}.xlsx")
        wb.save(file_path)
        return file_path, TreeExporter._build_preview(tree, preview_steps)


class TreePlotter:
    """Classe responsable des visualisations de l'arbre et des analyses (ultra-optimisée)."""

    # ------------------------------
    # Helpers optimisés
    # ------------------------------
    @staticmethod
    def _process_node_connections(node, x_pos, y_pos, step, max_steps, edges_x, edges_y, next_level):
        """Traite les connexions d'un nœud vers le niveau suivant."""
        if step < max_steps:
            for next_node, y_offset in [(node.next_up, 1), (node.next_mid, 0), (node.next_down, -1)]:
                if next_node:
                    next_y = y_pos + y_offset * 2
                    edges_x.extend([x_pos, x_pos + 1, None])
                    edges_y.extend([y_pos, next_y, None])
                    if next_node not in next_level:
                        next_level.append(next_node)

    @staticmethod
    def _extract_tree_data_fast(tree, max_steps):
        """Extraction ultra-rapide des données d'arbre pour Plotly."""
        positions, labels, edges_x, edges_y = [], [], [], []
        current_level = [tree.root]
        
        for step in range(min(max_steps + 1, tree.params.nb_steps + 1)):
            next_level = []
            level_size = len(current_level)
            
            for i, node in enumerate(current_level):
                if node is None:
                    continue
                y_pos = (i - level_size // 2) * 2
                x_pos = step
                
                positions.append((x_pos, y_pos))
                labels.append(f"{node.spot:.2f}")
                
                TreePlotter._process_node_connections(node, x_pos, y_pos, step, max_steps, edges_x, edges_y, next_level)
            
            current_level = next_level
            
        return positions, labels, edges_x, edges_y

    @staticmethod
    def _add_tree_edges(fig, edges_x, edges_y):
        """Ajoute les arêtes au graphique."""
        fig.add_trace(go.Scatter(
            x=edges_x, y=edges_y,
            mode='lines',
            line={'color': 'lightgray', 'width': 1},
            showlegend=False,
            hoverinfo='skip'
        ))

    @staticmethod
    def _add_tree_nodes(fig, positions, labels):
        """Ajoute les nœuds avec labels au graphique."""
        x_coords, y_coords = zip(*positions) if positions else ([], [])
        fig.add_trace(go.Scatter(
            x=x_coords, y=y_coords,
            mode='markers+text',
            marker={'size': 20, 'color': 'lightblue', 'line': {'width': 1, 'color': 'navy'}},
            text=labels,
            textposition='middle center',
            textfont={'size': 10, 'color': 'black'},
            showlegend=False,
            hovertemplate='Spot: %{text}<br>Step: %{x}<extra></extra>'
        ))

    @staticmethod
    def _create_plotly_tree(positions, labels, edges_x, edges_y, max_steps):
        """Crée un graphique Plotly interactif ultra-rapide."""
        fig = go.Figure()
        
        TreePlotter._add_tree_edges(fig, edges_x, edges_y)
        TreePlotter._add_tree_nodes(fig, positions, labels)
        
        fig.update_layout(
            title=f"Arbre trinomial (spots, max_steps={max_steps})",
            xaxis={'title': "Étapes", 'showgrid': True, 'zeroline': False},
            yaxis={'title': "Niveaux", 'showgrid': True, 'zeroline': True},
            showlegend=False,
            hovermode='closest',
            template='plotly_white',
            margin={'l': 50, 'r': 50, 't': 50, 'b': 50}
        )
        
        return fig

    # --------
    # Public 
    # --------
    @staticmethod
    def plot_tree(tree, max_steps=3, show_values=False):
        """Affiche l'arbre trinomial avec Plotly (ultra-rapide et interactif)."""
        if tree.last is None:
            tree.build_tree()
            tree.price_backward()
        
        # Extraction rapide des données
        positions, labels, edges_x, edges_y = TreePlotter._extract_tree_data_fast(tree, max_steps)
        
        # Création du graphique Plotly optimisé
        fig = TreePlotter._create_plotly_tree(positions, labels, edges_x, edges_y, max_steps)
        
        # Affichage avec optimisations Streamlit
        st.plotly_chart(fig, use_container_width=True, config={
            'displayModeBar': True,
            'displaylogo': False,
            'modeBarButtonsToRemove': ['pan2d', 'lasso2d', 'select2d']
        })

    @staticmethod
    def _optimize_convergence_data(steps, diffs):
        """Optimise les données de convergence pour l'affichage."""
        if len(steps) > 1000:
            stride = len(steps) // 500
            return steps[::stride], diffs[::stride]
        return steps, diffs

    @staticmethod
    def plot_convergence(diffs: np.ndarray, steps: np.ndarray, call_put: str):
        """Graphique de convergence optimisé pour les performances."""
        fig = go.Figure()
        
        steps_opt, diffs_opt = TreePlotter._optimize_convergence_data(steps, diffs)
        
        fig.add_trace(go.Scatter(
            x=steps_opt, y=diffs_opt, mode="lines+markers", name="(Tree - BS) × n",
            line={'width': 2}, marker={'size': 4}
        ))
        
        fig.add_hline(y=0, line_dash="dash", line_color="red", line_width=1)
        
        fig.update_layout(
            title=f"Convergence vers Black-Scholes ({call_put})",
            xaxis_title="Steps", yaxis_title="Erreur × n", template="plotly_white",
            showlegend=False, margin={'l': 50, 'r': 50, 't': 50, 'b': 50}
        )
        
        st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False, 'staticPlot': False})

    @staticmethod
    def _add_price_traces(fig, strikes, bs_prices, tree_prices):
        """Ajoute les traces de prix au graphique."""
        fig.add_trace(go.Scatter(
            x=strikes, y=bs_prices, mode="lines+markers", name="Prix BS",
            line={'width': 2, 'color': 'blue'}, marker={'size': 4}
        ), row=1, col=1)
        
        fig.add_trace(go.Scatter(
            x=strikes, y=tree_prices, mode="lines+markers", name="Prix Tree",
            line={'width': 2, 'color': 'red'}, marker={'size': 4}
        ), row=1, col=1)

    @staticmethod
    def _add_slope_traces(fig, strikes, bs_slopes, tree_slopes, diffs):
        """Ajoute les traces de pentes et différences."""
        fig.add_trace(go.Scatter(
            x=strikes[1:], y=bs_slopes, mode="lines", name="Pente BS",
            line={'width': 2, 'color': 'blue', 'dash': 'dot'}
        ), row=2, col=1)
        
        fig.add_trace(go.Scatter(
            x=strikes[1:], y=tree_slopes, mode="lines", name="Pente Tree",
            line={'width': 2, 'color': 'red', 'dash': 'dot'}
        ), row=2, col=1)
        
        fig.add_trace(go.Scatter(
            x=strikes, y=diffs, mode="lines+markers", name="Diff (Tree - BS)",
            line={'width': 2, 'color': 'green'}, marker={'size': 3}
        ), row=2, col=1)

    @staticmethod
    def plot_prices_and_slopes(strikes, bs_prices, tree_prices, bs_slopes, tree_slopes, diffs):
        """Graphique prix/pentes optimisé pour les performances."""
        fig = make_subplots(
            rows=2, cols=1, shared_xaxes=True,
            subplot_titles=("Comparaison des prix", "Pentes et différences"),
            vertical_spacing=0.1
        )
        
        TreePlotter._add_price_traces(fig, strikes, bs_prices, tree_prices)
        TreePlotter._add_slope_traces(fig, strikes, bs_slopes, tree_slopes, diffs)
        
        fig.update_layout(
            height=700, template="plotly_white",
            margin={'l': 50, 'r': 50, 't': 80, 'b': 50}, showlegend=True,
            legend={'orientation': "h", 'yanchor': "bottom", 'y': 1.02, 'xanchor': "right", 'x': 1}
        )
        
        st.plotly_chart(fig, use_container_width=True, config={
            'displayModeBar': True,
            'displaylogo': False,
            'modeBarButtonsToRemove': ['pan2d', 'lasso2d']
        })

    @staticmethod
    def plot_execution_times(times, method="backward"):
        """Graphique des temps d'exécution optimisé."""
        steps = list(range(1, len(times) + 1))
        
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=steps, y=times, mode="lines+markers", name=method,
            line={'width': 2}, marker={'size': 5},
            hovertemplate='Steps: %{x}<br>Temps: %{y:.4f}s<extra></extra>'
        ))
        
        fig.update_layout(
            title=f"Temps d'exécution ({method})",
            xaxis_title="Nombre de pas", yaxis_title="Temps (sec)",
            template="plotly_white", showlegend=False,
            margin={'l': 50, 'r': 50, 't': 50, 'b': 50}
        )
        
        st.plotly_chart(fig, use_container_width=True, config={
            'displayModeBar': False, 'staticPlot': False
        })


class MonteCarloPlotter:
    """Classe responsable des graphiques Monte Carlo."""

    @staticmethod
    def plot_convergence(mc_pricer, inputs):
        """Graphique de convergence Monte Carlo."""
        try:
            import plotly.express as px
            sim_points, prices = mc_pricer.convergence_study()
            
            fig = px.line(x=sim_points, y=prices, 
                         title="Convergence du prix Monte Carlo",
                         labels={"x": "Nombre de simulations", "y": "Prix estimé"})
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
        except Exception as e:
            st.error(f"Erreur lors du graphique de convergence : {e}")

    @staticmethod
    def plot_price_distribution(mc_pricer):
        """Histogramme de la distribution des prix finaux."""
        try:
            import plotly.express as px
            ST, _ = mc_pricer.get_simulation_data()
            
            fig = px.histogram(x=ST, nbins=50,
                              title="Distribution des prix finaux du sous-jacent",
                              labels={"x": "Prix final (ST)", "y": "Fréquence"})
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
            
            # Statistiques
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Prix moyen", f"{np.mean(ST):.2f}")
            with col2:
                st.metric("Écart-type", f"{np.std(ST):.2f}")
            with col3:
                st.metric("Médiane", f"{np.median(ST):.2f}")
        except Exception as e:
            st.error(f"Erreur lors du graphique de distribution des prix : {e}")

    @staticmethod
    def plot_payoff_distribution(mc_pricer):
        """Histogramme de la distribution des payoffs."""
        try:
            import plotly.express as px
            _, payoffs = mc_pricer.get_simulation_data()
            
            fig = px.histogram(x=payoffs, nbins=50,
                              title=f"Distribution des payoffs ({mc_pricer.option.call_put.upper()})",
                              labels={"x": "Payoff", "y": "Fréquence"})
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
            
            # Statistiques simplifiées
            positive_payoffs = payoffs[payoffs > 0]
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Payoff moyen", f"{np.mean(payoffs):.4f}")
            with col2:
                st.metric("% ITM", f"{len(positive_payoffs)/len(payoffs)*100:.1f}%")
            with col3:
                st.metric("Payoff max", f"{np.max(payoffs):.2f}")
        except Exception as e:
            st.error(f"Erreur lors du graphique de distribution des payoffs : {e}")


class BlackScholesPlotter:
    """Classe responsable des graphiques Black-Scholes."""

    @staticmethod
    def plot_sensitivity_analysis(bs_pricer):
        """Graphiques de sensibilité des Greeks."""
        try:
            import plotly.graph_objects as go
            from plotly.subplots import make_subplots
            
            data = bs_pricer.sensitivity_analysis()
            
            # Créer subplots 2x2
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=('Delta vs Spot', 'Gamma vs Spot', 'Vega vs Volatilité', 'Prix vs Temps'),
                specs=[[{"secondary_y": False}, {"secondary_y": False}],
                       [{"secondary_y": False}, {"secondary_y": False}]]
            )
            
            # Delta vs Spot
            fig.add_trace(go.Scatter(x=data['S_range'], y=data['deltas'], name='Delta'), row=1, col=1)
            
            # Gamma vs Spot
            fig.add_trace(go.Scatter(x=data['S_range'], y=data['gammas'], name='Gamma'), row=1, col=2)
            
            # Vega vs Volatilité
            fig.add_trace(go.Scatter(x=data['vol_range'], y=data['vegas'], name='Vega'), row=2, col=1)
            
            # Prix vs Temps
            fig.add_trace(go.Scatter(x=data['time_range'], y=data['prices_time'], name='Prix'), row=2, col=2)
            
            fig.update_layout(height=600, title_text="Analyse de sensibilité des Greeks")
            st.plotly_chart(fig, use_container_width=True)
            
        except Exception as e:
            st.error(f"Erreur lors de l'analyse de sensibilité : {e}")

    @staticmethod
    def plot_price_surface(bs_pricer):
        """Surface 3D des prix."""
        try:
            import plotly.graph_objects as go
            
            K_range, T_range, prices = bs_pricer.price_surface()
            
            fig = go.Figure(data=[go.Surface(z=prices, x=K_range, y=T_range)])
            fig.update_layout(
                title='Surface de prix Black-Scholes',
                scene=dict(
                    xaxis_title='Strike',
                    yaxis_title='Temps jusqu\'à maturité',
                    zaxis_title='Prix'
                ),
                height=500
            )
            st.plotly_chart(fig, use_container_width=True)
            
        except Exception as e:
            st.error(f"Erreur lors de la surface de prix : {e}")

    @staticmethod
    def _calculate_vol_smile(strikes, market):
        """Calcule le smile de volatilité simulé."""
        vol_smile = []
        for k in strikes:
            moneyness = k / market.S0
            vol_implied = market.vol * (1 + 0.1 * (moneyness - 1)**2)
            vol_smile.append(vol_implied)
        return vol_smile

    @staticmethod
    def plot_volatility_analysis(bs_pricer, market, option):
        """Analyse de la volatilité (smile)."""
        try:
            import plotly.graph_objects as go
            
            strikes = np.linspace(option.K * 0.7, option.K * 1.3, 20)
            vol_smile = BlackScholesPlotter._calculate_vol_smile(strikes, market)
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=strikes, y=vol_smile, name='Volatilité implicite', line={'color': 'blue'}))
            fig.add_hline(y=market.vol, line_dash="dash", line_color="red", annotation_text="Vol réalisée")
            
            fig.update_layout(
                title='Smile de volatilité (simulé)', xaxis_title='Strike',
                yaxis_title='Volatilité implicite', height=400
            )
            st.plotly_chart(fig, use_container_width=True)
            
            st.info("📝 **Note** : Ce smile est simulé pour démonstration. En réalité, on calculerait la volatilité implicite à partir des prix de marché.")
            
        except Exception as e:
            st.error(f"Erreur lors de l'analyse de volatilité : {e}")

    @staticmethod
    def _create_comparison_chart(comparison):
        """Crée le graphique de comparaison BS vs MC."""
        import plotly.graph_objects as go
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=['Black-Scholes', 'Monte Carlo'],
            y=[comparison['bs_price'], comparison['mc_price']],
            error_y={'type': 'data', 'array': [0, comparison['mc_stderr']]},
            name='Prix'
        ))
        
        fig.update_layout(
            title='Comparaison Black-Scholes vs Monte Carlo',
            yaxis_title='Prix', height=400
        )
        return fig

    @staticmethod
    def _display_comparison_metrics(comparison):
        """Affiche les métriques de comparaison."""
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Prix BS", f"{comparison['bs_price']:.6f}")
        with col2:
            st.metric("Prix MC", f"{comparison['mc_price']:.6f}")
        with col3:
            st.metric("Différence", f"{comparison['difference']:.6f}")
        with col4:
            st.metric("Erreur relative", f"{comparison['relative_error']:.3f}%")

    @staticmethod
    def plot_monte_carlo_comparison(bs_pricer, inputs):
        """Comparaison Black-Scholes vs Monte Carlo."""
        try:
            comparison = bs_pricer.compare_with_monte_carlo(inputs.get("n_paths", 10000), inputs.get("seed", 42))
            
            fig = BlackScholesPlotter._create_comparison_chart(comparison)
            st.plotly_chart(fig, use_container_width=True)
            
            BlackScholesPlotter._display_comparison_metrics(comparison)
            
            mc_lower = comparison['mc_price'] - 1.96 * comparison['mc_stderr']
            mc_upper = comparison['mc_price'] + 1.96 * comparison['mc_stderr']
            
            if mc_lower <= comparison['bs_price'] <= mc_upper:
                st.success("✅ Le prix Black-Scholes est dans l'intervalle de confiance à 95% de Monte Carlo")
            else:
                st.warning("⚠️ Le prix Black-Scholes est hors de l'intervalle de confiance à 95% de Monte Carlo")
                
        except Exception as e:
            st.error(f"Erreur lors de la comparaison BS vs MC : {e}")